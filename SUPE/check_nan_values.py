#!/usr/bin/env python3
"""
Simple script to check for 'nan' values in an Excel file.
"""

import os
import pandas as pd
import openpyxl

def check_with_pandas():
    """Check for 'nan' strings using pandas."""
    print("Checking with pandas...")
    fname = os.path.join(os.path.expanduser('~'), 'OneDrive', 'Documents', 'Python', 'SAP_Audit_Report.xlsx')
    print(f"Checking Excel file: {fname}")
    
    # Read all sheets
    xls = pd.ExcelFile(fname)
    for sheet_name in xls.sheet_names:
        print(f"\nChecking sheet: {sheet_name}")
        df = pd.read_excel(fname, sheet_name=sheet_name)
        
        # Check each column for 'nan' strings
        print("Columns with 'nan' values:")
        nan_found = False
        for col in df.columns:
            # Convert to string and count 'nan' literals
            count = df[col].astype(str).str.contains('nan').sum()
            if count > 0:
                print(f"  {col}: {count} 'nan' strings")
                nan_found = True
        
        if not nan_found:
            print("  None found")

def check_with_openpyxl():
    """Check for 'nan' strings using openpyxl (direct cell access)."""
    print("\nChecking with openpyxl (direct cell access)...")
    fname = os.path.join(os.path.expanduser('~'), 'OneDrive', 'Documents', 'Python', 'SAP_Audit_Report.xlsx')
    
    try:
        # Load workbook
        workbook = openpyxl.load_workbook(fname)
        
        # Check each sheet
        for sheet_name in workbook.sheetnames:
            print(f"\nChecking sheet: {sheet_name}")
            sheet = workbook[sheet_name]
            
            # Count cells with 'nan' value
            nan_count = 0
            
            # Iterate through all rows (skip header)
            for row in sheet.iter_rows(min_row=2):
                for cell in row:
                    if cell.value == 'nan':
                        nan_count += 1
            
            if nan_count > 0:
                print(f"Found {nan_count} 'nan' strings in sheet {sheet_name}")
            else:
                print(f"No 'nan' strings found in sheet {sheet_name}")
                
    except Exception as e:
        print(f"Error checking with openpyxl: {str(e)}")

def main():
    """Main function."""
    print("=== Excel NaN Value Checker ===\n")
    
    # Check using pandas (detects 'nan' as a substring)
    check_with_pandas()
    
    # Check using openpyxl (direct cell access, only finds exact 'nan' matches)
    try:
        import openpyxl
        check_with_openpyxl()
    except ImportError:
        print("\nopenpyxl not installed. Skipping direct cell check.")
    
    print("\nInspection complete.")

if __name__ == "__main__":
    main()
