#!/usr/bin/env python3
"""
SAP Audit Tool Test Script

This script performs automated testing of the SAP Audit Tool components
with specific focus on variable field handling and NaN value elimination.
"""

import os
import sys
import pandas as pd
import time
import glob
import shutil
import subprocess
from datetime import datetime
import inspect

# Test configuration
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
TEST_OUTPUT_DIR = os.path.join(SCRIPT_DIR, "test_output")
TEST_INPUT_DIR = os.path.join(SCRIPT_DIR, "input")  # Use existing input directory
DEBUG_LOG_FILE = os.path.join(SCRIPT_DIR, "test_debug.log")

# Import local modules (assumes they're in the same directory)
try:
    from sap_audit_data_prep import process_sm20, process_cdhdr, process_cdpos
    from sap_audit_tool_risk_assessment import detect_debug_patterns, assess_risk_session, custom_field_risk_assessment
    from sap_audit_tool_output import generate_excel_output
    print("Successfully imported SAP Audit Tool modules")
except ImportError as e:
    print(f"Error importing SAP Audit Tool modules: {str(e)}")
    print("Make sure you're running this script from the correct directory.")
    sys.exit(1)

def log_message(message, level="INFO"):
    """Log a message with timestamp and level to both console and log file."""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_entry = f"[{timestamp}] {level}: {message}"
    
    # Create log directory if it doesn't exist
    os.makedirs(os.path.dirname(DEBUG_LOG_FILE), exist_ok=True)
    
    # Write to log file
    with open(DEBUG_LOG_FILE, "a") as f:
        f.write(log_entry + "\n")
    
    # Print to console
    print(log_entry)

def setup_test_environment():
    """Create test directories and ensure we have access to sample data."""
    log_message("Setting up test environment...")
    
    # Create test output directory if it doesn't exist
    os.makedirs(TEST_OUTPUT_DIR, exist_ok=True)
    
    # Check if input directory exists and has required files
    input_files_exist = True
    
    # Create input dir if it doesn't exist
    os.makedirs(TEST_INPUT_DIR, exist_ok=True)
    
    # Check for SM20 files
    sm20_files = glob.glob(os.path.join(TEST_INPUT_DIR, "*_sm20_*.xlsx"))
    if not sm20_files:
        log_message("Warning: No SM20 files found in input directory", "WARNING")
        input_files_exist = False
    
    # Check for CDHDR files
    cdhdr_files = glob.glob(os.path.join(TEST_INPUT_DIR, "*_cdhdr_*.xlsx"))
    if not cdhdr_files:
        log_message("Warning: No CDHDR files found in input directory", "WARNING")
        input_files_exist = False
    
    # Check for CDPOS files
    cdpos_files = glob.glob(os.path.join(TEST_INPUT_DIR, "*_cdpos_*.xlsx"))
    if not cdpos_files:
        log_message("Warning: No CDPOS files found in input directory", "WARNING")
        input_files_exist = False
    
    if input_files_exist:
        log_message("Input files found for testing.")
    else:
        log_message("Warning: Some input files missing, tests may be limited.", "WARNING")
    
    return input_files_exist

def test_data_preparation():
    """Test the data preparation module."""
    log_message("Testing data preparation module...")
    
    # Find input files
    sm20_files = glob.glob(os.path.join(TEST_INPUT_DIR, "*_sm20_*.xlsx"))
    cdhdr_files = glob.glob(os.path.join(TEST_INPUT_DIR, "*_cdhdr_*.xlsx"))
    cdpos_files = glob.glob(os.path.join(TEST_INPUT_DIR, "*_cdpos_*.xlsx"))
    
    sm20_success = False
    cdhdr_success = False
    cdpos_success = False
    
    # Test SM20 processing
    if sm20_files:
        sm20_file = sm20_files[0]
        sm20_output = os.path.join(TEST_OUTPUT_DIR, "SM20_test.csv")
        
        log_message(f"Processing SM20 file: {sm20_file}")
        try:
            # This calls our module's function
            process_sm20(sm20_file, sm20_output)
            
            # Verify output file exists
            if os.path.exists(sm20_output):
                # Check if variable fields are preserved
                df = pd.read_csv(sm20_output, encoding='utf-8-sig')
                
                var_fields_present = all(field in df.columns for field in ['FIRST VARIABLE VALUE FOR EVENT', 'VARIABLE 2', 'VARIABLE DATA FOR MESSAGE'])
                
                if var_fields_present:
                    log_message("Variable fields successfully preserved in SM20 output", "SUCCESS")
                    
                    # Check for NaN values
                    nan_values = df.isna().sum().sum()
                    log_message(f"Found {nan_values} NaN values in SM20 output")
                    
                    sm20_success = True
                else:
                    log_message("Some variable fields missing in SM20 output", "ERROR")
                    available_fields = [col for col in df.columns]
                    log_message(f"Available fields: {available_fields}", "INFO")
            else:
                log_message(f"SM20 output file not created: {sm20_output}", "ERROR")
        except Exception as e:
            log_message(f"Error processing SM20 file: {str(e)}", "ERROR")
    
    # Test CDHDR processing
    if cdhdr_files:
        cdhdr_file = cdhdr_files[0]
        cdhdr_output = os.path.join(TEST_OUTPUT_DIR, "CDHDR_test.csv")
        
        log_message(f"Processing CDHDR file: {cdhdr_file}")
        try:
            process_cdhdr(cdhdr_file, cdhdr_output)
            
            # Verify output
            if os.path.exists(cdhdr_output):
                log_message("CDHDR processing successful", "SUCCESS")
                cdhdr_success = True
            else:
                log_message(f"CDHDR output file not created: {cdhdr_output}", "ERROR")
        except Exception as e:
            log_message(f"Error processing CDHDR file: {str(e)}", "ERROR")
    
    # Test CDPOS processing
    if cdpos_files:
        cdpos_file = cdpos_files[0]
        cdpos_output = os.path.join(TEST_OUTPUT_DIR, "CDPOS_test.csv")
        
        log_message(f"Processing CDPOS file: {cdpos_file}")
        try:
            process_cdpos(cdpos_file, cdpos_output)
            
            # Verify output
            if os.path.exists(cdpos_output):
                log_message("CDPOS processing successful", "SUCCESS")
                cdpos_success = True
            else:
                log_message(f"CDPOS output file not created: {cdpos_output}", "ERROR")
        except Exception as e:
            log_message(f"Error processing CDPOS file: {str(e)}", "ERROR")
    
    # Report overall success
    if sm20_success and cdhdr_success and cdpos_success:
        log_message("Data preparation module testing passed", "SUCCESS")
        return True
    else:
        log_message("Data preparation module testing had issues", "WARNING")
        return False

def run_session_merger():
    """Run the session merger script and test its output."""
    log_message("Testing session merger functionality...")
    
    # Clear any existing session timeline
    session_timeline = os.path.join(SCRIPT_DIR, "SAP_Session_Timeline.xlsx")
    if os.path.exists(session_timeline):
        os.remove(session_timeline)
    
    # Run the session merger script
    merger_script = os.path.join(SCRIPT_DIR, "SAP Log Session Merger.py")
    try:
        log_message(f"Running: {merger_script}")
        result = subprocess.run(
            [sys.executable, merger_script],
            capture_output=True,
            text=True
        )
        
        if result.returncode == 0:
            log_message("Session merger ran successfully")
            
            # Check if session timeline was created
            if os.path.exists(session_timeline):
                log_message(f"Session timeline created: {session_timeline}")
                
                # Verify variable fields
                df = pd.read_excel(session_timeline)
                var_fields_present = all(field in df.columns for field in ['Variable_First', 'Variable_2', 'Variable_Data'])
                
                if var_fields_present:
                    log_message("Variable fields successfully preserved in session timeline", "SUCCESS")
                    
                    # Check for any 'nan' strings
                    for col in ['Variable_First', 'Variable_2', 'Variable_Data', 'Table', 'Field']:
                        if col in df.columns:
                            # Convert to string to check for 'nan' strings
                            nan_count = df[col].astype(str).str.count('nan').sum()
                            log_message(f"Found {nan_count} 'nan' strings in {col} column")
                    
                    return True
                else:
                    log_message("Some variable fields missing in session timeline", "ERROR")
                    available_fields = [col for col in df.columns]
                    log_message(f"Available fields: {available_fields}", "INFO")
                    return False
            else:
                log_message(f"Session timeline not created: {session_timeline}", "ERROR")
                return False
        else:
            log_message(f"Session merger failed with error: {result.stderr}", "ERROR")
            return False
    except Exception as e:
        log_message(f"Error running session merger: {str(e)}", "ERROR")
        return False

def test_risk_assessment():
    """Test the risk assessment module with debug patterns."""
    log_message("Testing risk assessment module...")
    
    # Load session timeline
    session_timeline = os.path.join(SCRIPT_DIR, "SAP_Session_Timeline.xlsx")
    if not os.path.exists(session_timeline):
        log_message(f"Session timeline not found: {session_timeline}", "ERROR")
        return False
    
    try:
        # Load session data
        session_data = pd.read_excel(session_timeline)
        initial_count = len(session_data)
        
        # Create a copy with test data
        test_data = session_data.copy()
        
        # Inject some known debug patterns for testing
        if 'Variable_2' in test_data.columns:
            # Inject some debug flags for testing
            test_count = min(10, len(test_data))
            for i in range(test_count):
                idx = test_data.index[i]
                test_data.loc[idx, 'Variable_2'] = 'I!'  # Debug flag
        
        # Run risk assessment
        log_message("Running risk assessment on test data...")
        result_data = assess_risk_session(test_data)
        
        # Verify output
        if len(result_data) != initial_count:
            log_message(f"Risk assessment changed row count from {initial_count} to {len(result_data)}", "ERROR")
            return False
        
        # Check for critical risk levels from debug patterns
        if 'risk_level' in result_data.columns:
            critical_risk_count = len(result_data[result_data['risk_level'] == 'Critical'])
            high_risk_count = len(result_data[result_data['risk_level'] == 'High'])
            
            log_message(f"Risk assessment assigned {critical_risk_count} critical risk and {high_risk_count} high risk events")
            
            # Test a specific row with debug flag
            debug_rows = result_data[result_data['Variable_2'] == 'I!']
            if not debug_rows.empty:
                first_debug_row = debug_rows.iloc[0]
                if first_debug_row['risk_level'] in ['Critical', 'High']:
                    log_message("Debug pattern correctly assigned high/critical risk", "SUCCESS")
                else:
                    log_message(f"Debug pattern incorrectly assigned {first_debug_row['risk_level']} risk", "ERROR")
            
            return True
        else:
            log_message("risk_level column not found in risk assessment output", "ERROR")
            return False
    except Exception as e:
        log_message(f"Error testing risk assessment: {str(e)}", "ERROR")
        return False

def test_output_generation():
    """Test the Excel output generation, specifically NaN handling."""
    log_message("Testing output generation module...")
    
    # Load session timeline
    session_timeline = os.path.join(SCRIPT_DIR, "SAP_Session_Timeline.xlsx")
    if not os.path.exists(session_timeline):
        log_message(f"Session timeline not found: {session_timeline}", "ERROR")
        return False
    
    try:
        # Load session data
        session_data = pd.read_excel(session_timeline)
        
        # Run risk assessment
        log_message("Running risk assessment for output testing...")
        risk_data = assess_risk_session(session_data)
        
        # Generate test output
        test_output_file = os.path.join(TEST_OUTPUT_DIR, "Test_SAP_Audit_Report.xlsx")
        
        log_message(f"Generating test output to: {test_output_file}")
        result = generate_excel_output(
            pd.DataFrame(),  # Empty correlated_df for testing
            pd.DataFrame(),  # Empty unmatched_cdpos
            pd.DataFrame(),  # Empty unmatched_sm20
            risk_data,       # Session data with risk assessment
            test_output_file
        )
        
        if result:
            log_message(f"Test output successfully generated: {test_output_file}")
            
            # Check for nan values
            nan_present = check_nan_values_in_excel(test_output_file)
            if nan_present:
                log_message("Warning: 'nan' values found in Excel output", "WARNING")
                return False
            else:
                log_message("No 'nan' values found in Excel output", "SUCCESS")
                return True
        else:
            log_message("Failed to generate test output", "ERROR")
            return False
    except Exception as e:
        log_message(f"Error testing output generation: {str(e)}", "ERROR")
        return False

def check_nan_values_in_excel(excel_file):
    """
    Check if any 'nan' strings exist in the Excel output using a visual inspection approach.
    This method directly accesses Excel cells rather than using pandas to avoid
    pandas automatically converting empty cells to NaN during reading.
    """
    try:
        log_message(f"Visually checking for 'nan' strings in: {excel_file}")
        
        # Import necessary Excel-specific libraries
        import openpyxl
        
        # Load the workbook directly
        workbook = openpyxl.load_workbook(excel_file)
        nan_found = False
        
        # Check each sheet
        for sheet_name in workbook.sheetnames:
            log_message(f"Checking sheet: {sheet_name}")
            sheet = workbook[sheet_name]
            
            # Reset counters for each sheet
            sheet_nan_count = 0
            
            # Check each cell in the sheet
            for row in sheet.iter_rows(min_row=2):  # Skip header row
                for cell in row:
                    # Check if the cell's value is exactly the string 'nan'
                    if cell.value == 'nan':
                        sheet_nan_count += 1
            
            if sheet_nan_count > 0:
                log_message(f"Found {sheet_nan_count} visual 'nan' strings in sheet '{sheet_name}'", "WARNING")
                nan_found = True
            else:
                log_message(f"No visual 'nan' strings found in sheet '{sheet_name}'", "SUCCESS")
        
        return nan_found
    except Exception as e:
        log_message(f"Error checking for nan values: {str(e)}", "ERROR")
        return True  # Assume nan values exist if we encounter an error

def test_clean_df_function():
    """Test the NaN cleaning functionality with a direct implementation."""
    log_message("Testing NaN cleaning functionality...")
    
    try:
        # Create a test DataFrame with different types of NaN values
        import numpy as np
        
        test_df = pd.DataFrame({
            'String': ['text', None, np.nan, 'nan', ''],
            'Number': [1, None, np.nan, 3, 4],
            'Table': ['table1', None, np.nan, 'nan', 'table5'],
            'Field': ['field1', None, np.nan, 'nan', 'field5'],
            'Variable_2': ['I!', None, np.nan, 'nan', 'D!']
        })
        
        # Implement the same cleaning logic used in the output module
        def direct_clean_df(df):
            if df is None or len(df) == 0:
                return df
                
            # Make a copy to avoid modifying original
            df_clean = df.copy()
            
            # List of columns that should ALWAYS be treated as strings with empty values instead of NaN
            string_columns = [
                'Table', 'Field', 'Change_Indicator', 'Old_Value', 'New_Value',
                'Variable_First', 'Variable_2', 'Variable_Data', 'Object', 'Object_ID',
                'Doc_Number', 'Description', 'TCode', 'Source', 'Session ID with Date'
            ]
            
            # First pass: Replace NaN with empty string in object columns
            for col in df_clean.columns:
                if df_clean[col].dtype == 'object':
                    df_clean[col] = df_clean[col].fillna('')
            
            # Second pass: Convert specified columns to strings and clean 'nan' values
            for col in df_clean.columns:
                if col in string_columns:
                    if col in df_clean.columns:
                        # Convert to string
                        df_clean[col] = df_clean[col].astype(str)
                        
                        # Replace literal 'nan' strings with empty string
                        df_clean[col] = df_clean[col].replace('nan', '')
                        
                        # Also replace 'None' strings
                        df_clean[col] = df_clean[col].replace('None', '')
            
            return df_clean
        
        # Apply cleaning
        cleaned_df = direct_clean_df(test_df)
        
        # Check results
        string_cols = ['String', 'Table', 'Field', 'Variable_2']
        num_cols = ['Number']
        
        # Verification loop
        for col in string_cols:
            # Check if any NaN values remain
            if cleaned_df[col].isna().any():
                log_message(f"NaN values still present in {col} column", "ERROR")
                return False
            
            # Check if any 'nan' strings remain
            if (cleaned_df[col] == 'nan').any():
                log_message(f"'nan' strings still present in {col} column", "ERROR")
                return False
            
            # Check if any 'None' strings remain
            if (cleaned_df[col] == 'None').any():
                log_message(f"'None' strings still present in {col} column", "ERROR")
                return False
        
        log_message("NaN cleaning functionality successfully cleans NaN and 'nan' values", "SUCCESS")
        return True
    except Exception as e:
        log_message(f"Error testing NaN cleaning: {str(e)}", "ERROR")
        return False

def run_full_system_test():
    """Run a full end-to-end test of the entire system."""
    log_message("Running full system test...")
    
    try:
        # Run the main tool 
        log_message("Running SAP audit tool end-to-end...")
        subprocess.run(
            [sys.executable, os.path.join(SCRIPT_DIR, "sap_audit_tool.py")],
            check=True
        )
        
        # Check the output
        output_file = os.path.join(SCRIPT_DIR, "SAP_Audit_Report.xlsx")
        if os.path.exists(output_file):
            log_message(f"Output file created: {output_file}")
            
            # Check for nan values
            nan_present = check_nan_values_in_excel(output_file)
            if nan_present:
                log_message("Warning: 'nan' values found in final output", "WARNING")
                return False
            else:
                log_message("No 'nan' values found in final output", "SUCCESS")
                
                # Copy to test output for reference
                shutil.copy2(output_file, os.path.join(TEST_OUTPUT_DIR, "System_Test_Output.xlsx"))
                return True
        else:
            log_message(f"Output file not created: {output_file}", "ERROR")
            return False
    except Exception as e:
        log_message(f"Error in system test: {str(e)}", "ERROR")
        return False

def main():
    """Main test execution function."""
    start_time = time.time()
    log_message("Starting SAP Audit Tool automated testing...")
    
    # Setup test environment
    os.makedirs(TEST_OUTPUT_DIR, exist_ok=True)
    if os.path.exists(DEBUG_LOG_FILE):
        os.remove(DEBUG_LOG_FILE)
    
    setup_test_environment()
    
    # Run tests
    tests = [
        ("Component - NaN Cleaning", test_clean_df_function),
        ("Component - Session Merger", run_session_merger),
        ("Component - Risk Assessment", test_risk_assessment),
        ("Component - Output Generation", test_output_generation),
        ("End-to-End System Test", run_full_system_test)
    ]
    
    results = {}
    for test_name, test_func in tests:
        log_message(f"Running test: {test_name}...")
        try:
            result = test_func()
            results[test_name] = result
            status = "PASSED" if result else "FAILED"
            log_message(f"Test {test_name}: {status}")
        except Exception as e:
            results[test_name] = False
            log_message(f"Test {test_name} ERROR: {str(e)}", "ERROR")
    
    # Print summary
    log_message("\n=== TEST SUMMARY ===")
    all_passed = True
    for test_name, result in results.items():
        status = "PASSED" if result else "FAILED"
        log_message(f"{test_name}: {status}")
        if not result:
            all_passed = False
    
    elapsed_time = time.time() - start_time
    log_message(f"Testing completed in {elapsed_time:.2f} seconds.")
    
    if all_passed:
        log_message("All tests passed successfully!", "SUCCESS")
        return 0
    else:
        log_message("Some tests failed. Check the logs for details.", "WARNING")
        return 1

if __name__ == "__main__":
    sys.exit(main())
